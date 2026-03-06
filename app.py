from flask import Flask, render_template, request, jsonify, redirect, url_for, session, send_file, flash
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from datetime import datetime, timedelta, timezone, date
from functools import wraps
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import io
from flask import request, jsonify
from sqlalchemy import inspect, text
from zoneinfo import ZoneInfo

from models import db, User, Student, AttendanceLog, Bill, Payment, Complaint, Notification

app = Flask(__name__)
STAFF_ROLES = ('admin', 'warden', 'principal')
IST = ZoneInfo("Asia/Kolkata")

app.config["SECRET_KEY"] = os.environ.get(
    "SECRET_KEY",
    "dormio-secret-key-2024"
)

# Database connection from environment variable
app.config["SQLALCHEMY_DATABASE_URI"] = os.environ.get("DATABASE_URL")

app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

# Important for serverless platforms
app.config["SQLALCHEMY_ENGINE_OPTIONS"] = {
    "pool_pre_ping": True,
    "pool_recycle": 300
}

db.init_app(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# Role-based access control decorator
def role_required(*roles):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not current_user.is_authenticated or current_user.role not in roles:
                flash('Access denied. Insufficient permissions.', 'error')
                return redirect(url_for('dashboard'))
            return f(*args, **kwargs)
        return decorated_function
    return decorator

def get_staff_route(warden_endpoint, admin_endpoint):
    return admin_endpoint if current_user.role == 'admin' else warden_endpoint

def to_ist(dt_value):
    if not dt_value:
        return dt_value
    if dt_value.tzinfo is None:
        dt_value = dt_value.replace(tzinfo=timezone.utc)
    return dt_value.astimezone(IST)

@app.template_filter('format_datetime')
def format_datetime(value, fmt='%Y-%m-%d %H:%M:%S'):
    if not value:
        return ''
    if isinstance(value, datetime):
        value = to_ist(value)
    return value.strftime(fmt)

def ensure_notification_message_columns():
    inspector = inspect(db.engine)
    columns = {c['name'] for c in inspector.get_columns('notification')}

    alter_statements = []
    if 'sender_id' not in columns:
        alter_statements.append("ALTER TABLE notification ADD COLUMN sender_id INTEGER")
    if 'parent_notification_id' not in columns:
        alter_statements.append("ALTER TABLE notification ADD COLUMN parent_notification_id INTEGER")

    for stmt in alter_statements:
        db.session.execute(text(stmt))

    if alter_statements:
        db.session.commit()

# Initialize database with default users
with app.app_context():
    db.create_all()
    ensure_notification_message_columns()
    
    # Create default users if not exist
    if not User.query.filter_by(username='admin').first():
        admin = User(username='admin', email='admin@dormio.com', full_name='System Administrator', role='admin')
        admin.set_password('admin123')
        db.session.add(admin)
    
    if not User.query.filter_by(username='warden').first():
        warden = User(username='warden', email='warden@dormio.com', full_name='Hostel Warden', role='warden')
        warden.set_password('warden123')
        db.session.add(warden)
    
    if not User.query.filter_by(username='principal').first():
        principal = User(username='principal', email='principal@dormio.com', full_name='College Principal', role='principal')
        principal.set_password('principal123')
        db.session.add(principal)
    
    db.session.commit()
    print("Default users created:")
    print("Admin - username: admin, password: admin123")
    print("Warden - username: warden, password: warden123")
    print("Principal - username: principal, password: principal123")

# Routes
@app.route('/')
def index():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        user = User.query.filter_by(username=username).first()
        
        if user and user.check_password(password) and user.is_active:
            login_user(user)
            return redirect(url_for('dashboard'))
        else:
            flash('Invalid credentials or account disabled', 'error')
    
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route('/dashboard')
@login_required
def dashboard():
    # Route to appropriate dashboard based on role
    if current_user.role == 'student':
        return redirect(url_for('student_dashboard'))
    elif current_user.role == 'warden':
        return redirect(url_for('warden_dashboard'))
    elif current_user.role == 'principal':
        return redirect(url_for('principal_dashboard'))
    else:  # admin
        return redirect(url_for('admin_dashboard'))

# Student Module
@app.route('/student/dashboard')
@login_required
@role_required('student')
def student_dashboard():
    student = Student.query.filter_by(user_id=current_user.id).first()
    if not student:
        flash('Student profile not found', 'error')
        return redirect(url_for('logout'))
    
    # Get pending bills
    pending_bills = Bill.query.filter_by(student_id=student.id, paid=False).count()
    total_due = db.session.query(db.func.sum(Bill.amount)).filter_by(student_id=student.id, paid=False).scalar() or 0
    
    # Get overdue bills count
    today = date.today()
    overdue_count = Bill.query.filter_by(student_id=student.id, paid=False).filter(Bill.due_date < today).count()
    
    # Get recent complaints
    recent_complaints = Complaint.query.filter_by(student_id=student.id).order_by(Complaint.created_at.desc()).limit(5).all()
    
    # Get recent attendance
    recent_logs = AttendanceLog.query.filter_by(student_id=student.id).order_by(AttendanceLog.timestamp.desc()).limit(10).all()
    for log in recent_logs:
        log.timestamp = to_ist(log.timestamp)
    
    # Get unread notifications
    unread_notifications = Notification.query.filter_by(user_id=current_user.id, read=False).count()
    
    return render_template('student/dashboard.html',
                         student=student,
                         pending_bills=pending_bills,
                         total_due=total_due,
                         overdue_count=overdue_count,
                         recent_complaints=recent_complaints,
                         recent_logs=recent_logs,
                         unread_notifications=unread_notifications)

@app.route('/student/attendance')
@login_required
@role_required('student')
def student_attendance():
    student = Student.query.filter_by(user_id=current_user.id).first()
    logs = AttendanceLog.query.filter_by(student_id=student.id).order_by(AttendanceLog.timestamp.desc()).all()
    for log in logs:
        log.timestamp = to_ist(log.timestamp)
    return render_template('student/attendance.html', logs=logs, student=student)

@app.route('/student/bills')
@login_required
@role_required('student')
def student_bills():
    student = Student.query.filter_by(user_id=current_user.id).first()
    bills = Bill.query.filter_by(student_id=student.id).order_by(Bill.created_at.desc()).all()
    
    # Calculate overdue bills
    today = date.today()
    overdue_bills = [b for b in bills if not b.paid and b.due_date < today]
    
    # Payment statistics
    total_paid = sum(b.amount for b in bills if b.paid)
    total_pending = sum(b.amount for b in bills if not b.paid)
    
    return render_template('student/bills.html', 
                         bills=bills, 
                         student=student,
                         overdue_bills=overdue_bills,
                         total_paid=total_paid,
                         total_pending=total_pending,
                         today=today)

@app.route('/student/complaints', methods=['GET', 'POST'])
@login_required
@role_required('student')
def student_complaints():
    student = Student.query.filter_by(user_id=current_user.id).first()
    
    if request.method == 'POST':
        complaint = Complaint(
            student_id=student.id,
            ticket_id=Complaint.generate_ticket_id(),
            category=request.form.get('category'),
            subject=request.form.get('subject'),
            description=request.form.get('description'),
            priority=request.form.get('priority', 'medium')
        )
        db.session.add(complaint)
        
        # Create notification for wardens
        wardens = User.query.filter_by(role='warden').all()
        for warden in wardens:
            notification = Notification(
                user_id=warden.id,
                title='New Complaint Raised',
                message=f'{current_user.full_name} raised a complaint: {complaint.subject}',
                type='complaint'
            )
            db.session.add(notification)
        
        db.session.commit()
        flash(f'Complaint submitted successfully. Ticket ID: {complaint.ticket_id}', 'success')
        return redirect(url_for('student_complaints'))
    
    complaints = Complaint.query.filter_by(student_id=student.id).order_by(Complaint.created_at.desc()).all()
    return render_template('student/complaints.html', complaints=complaints, student=student)

# Warden Module
@app.route('/warden/dashboard')
@login_required
@role_required('warden')
def warden_dashboard():
    total_students = Student.query.count()
    students_in = Student.query.filter_by(current_status='IN').count()
    students_out = Student.query.filter_by(current_status='OUT').count()
    pending_complaints = Complaint.query.filter(Complaint.status.in_(['open', 'in_progress'])).count()
    
    # Recent activity
    recent_logs = db.session.query(AttendanceLog, Student)\
        .select_from(AttendanceLog)\
        .join(Student, AttendanceLog.student_id == Student.id)\
        .order_by(AttendanceLog.timestamp.desc())\
        .limit(10).all()
    for log, _student in recent_logs:
        log.timestamp = to_ist(log.timestamp)
    
    # Pending complaints
    recent_complaints = db.session.query(Complaint, Student)\
        .select_from(Complaint)\
        .join(Student, Complaint.student_id == Student.id)\
        .filter(Complaint.status.in_(['open', 'in_progress']))\
        .order_by(Complaint.created_at.desc())\
        .limit(5).all()
    
    return render_template('warden/dashboard.html',
                         total_students=total_students,
                         students_in=students_in,
                         students_out=students_out,
                         pending_complaints=pending_complaints,
                         recent_logs=recent_logs,
                         recent_complaints=recent_complaints)

@app.route('/warden/students', methods=['GET', 'POST'])
@app.route('/admin/students', methods=['GET', 'POST'], endpoint='admin_students')
@login_required
@role_required('warden', 'admin')
def warden_students():
    if request.method == 'POST':
        # Create user account
        username = request.form.get('username')
        password = request.form.get('password', 'student123')  # Default password if not provided
        
        # Check if username already exists
        if User.query.filter_by(username=username).first():
            flash('Username already exists', 'error')
            return redirect(url_for(get_staff_route('warden_students', 'admin_students')))
        
        user = User(
            username=username,
            email=request.form.get('email'),
            full_name=request.form.get('full_name'),
            role='student'
        )
        user.set_password(password)
        db.session.add(user)
        db.session.flush()
        
        # Check if RFID code already exists
        rfid_code = request.form.get('rfid_code')
        if Student.query.filter_by(rfid_code=rfid_code).first():
            db.session.rollback()
            flash('RFID code already exists', 'error')
            return redirect(url_for(get_staff_route('warden_students', 'admin_students')))
        
        # Create student profile
        student = Student(
            user_id=user.id,
            rfid_code=rfid_code,
            roll_number=request.form.get('roll_number'),
            room_number=request.form.get('room_number'),
            contact=request.form.get('contact'),
            semester=request.form.get('semester', 1),
            course=request.form.get('course', '')
        )
        db.session.add(student)
        
        # Create welcome notification
        notification = Notification(
            user_id=user.id,
            title='Welcome to Dormio',
            message=f'Your account has been created. Username: {username}. You can now login and access your hostel portal.',
            type='general'
        )
        db.session.add(notification)
        
        db.session.commit()
        flash(f'Student registered successfully. Username: {username}, Password: {password}', 'success')
        return redirect(url_for(get_staff_route('warden_students', 'admin_students')))
    
    students = db.session.query(Student, User)\
        .select_from(Student)\
        .join(User, Student.user_id == User.id)\
        .order_by(Student.created_at.desc()).all()
    template_name = 'admin/students.html' if current_user.role == 'admin' else 'warden/students.html'
    return render_template(template_name, students=students)

@app.route('/warden/students/delete/<int:id>', methods=['POST'])
@login_required
@role_required('warden', 'admin')
def delete_student(id):
    student = Student.query.get_or_404(id)
    user = User.query.get(student.user_id)
    db.session.delete(student)
    db.session.delete(user)
    db.session.commit()
    flash('Student deleted successfully', 'success')
    return redirect(url_for(get_staff_route('warden_students', 'admin_students')))

@app.route('/warden/bills', methods=['GET', 'POST'])
@app.route('/admin/bills', methods=['GET', 'POST'], endpoint='admin_bills')
@login_required
@role_required('warden', 'admin')
def warden_bills():
    if request.method == 'POST':
        student_id = request.form.get('student_id')
        bill = Bill(
            student_id=student_id,
            bill_type=request.form.get('bill_type'),
            amount=float(request.form.get('amount')),
            month=request.form.get('month'),
            semester=request.form.get('semester'),
            due_date=datetime.strptime(request.form.get('due_date'), '%Y-%m-%d').date(),
            created_by=current_user.id
        )
        db.session.add(bill)
        
        # Create notification for student
        student = Student.query.get(student_id)
        notification = Notification(
            user_id=student.user_id,
            title='New Bill Generated',
            message=f'A new {bill.bill_type} bill of ₹{bill.amount} has been generated for {bill.month}. Due date: {bill.due_date.strftime("%Y-%m-%d")}',
            type='bill'
        )
        db.session.add(notification)
        
        db.session.commit()
        flash('Bill created successfully', 'success')
        return redirect(url_for(get_staff_route('warden_bills', 'admin_bills')))
    
    bills = db.session.query(Bill, Student, User)\
        .select_from(Bill)\
        .join(Student, Bill.student_id == Student.id)\
        .join(User, Student.user_id == User.id)\
        .order_by(Bill.created_at.desc()).all()
    students = db.session.query(Student, User)\
        .select_from(Student)\
        .join(User, Student.user_id == User.id).all()
    template_name = 'admin/bills.html' if current_user.role == 'admin' else 'warden/bills.html'
    return render_template(template_name, bills=bills, students=students)

@app.route('/warden/bills/mark-paid/<int:id>', methods=['POST'])
@login_required
@role_required('warden', 'admin')
def mark_bill_paid(id):
    bill = Bill.query.get_or_404(id)
    bill.paid = True
    bill.paid_date = datetime.now(timezone.utc)
    
    # Create payment record
    payment = Payment(
        student_id=bill.student_id,
        bill_id=bill.id,
        amount=bill.amount,
        payment_method=request.form.get('payment_method', 'cash'),
        transaction_id=request.form.get('transaction_id'),
        remarks=request.form.get('remarks')
    )
    db.session.add(payment)
    
    # Create notification for student
    student = Student.query.get(bill.student_id)
    notification = Notification(
        user_id=student.user_id,
        title='Payment Confirmed',
        message=f'Your {bill.bill_type} bill for {bill.month} has been marked as paid. Amount: ₹{bill.amount}',
        type='bill'
    )
    db.session.add(notification)
    
    db.session.commit()
    flash('Bill marked as paid', 'success')
    return redirect(url_for(get_staff_route('warden_bills', 'admin_bills')))

@app.route('/warden/complaints')
@app.route('/admin/complaints', endpoint='admin_complaints')
@login_required
@role_required('warden', 'admin')
def warden_complaints():
    status_filter = request.args.get('status', 'all')
    query = db.session.query(Complaint, Student, User)\
        .select_from(Complaint)\
        .join(Student, Complaint.student_id == Student.id)\
        .join(User, Student.user_id == User.id)
    
    if status_filter != 'all':
        query = query.filter(Complaint.status == status_filter)
    
    complaints = query.order_by(Complaint.created_at.desc()).all()
    template_name = 'admin/complaints.html' if current_user.role == 'admin' else 'warden/complaints.html'
    return render_template(template_name, complaints=complaints, status_filter=status_filter)

@app.route('/warden/complaints/update/<int:id>', methods=['POST'])
@login_required
@role_required('warden', 'admin')
def update_complaint(id):
    complaint = Complaint.query.get_or_404(id)
    old_status = complaint.status
    complaint.status = request.form.get('status')
    complaint.resolution_notes = request.form.get('resolution_notes')
    complaint.assigned_to = current_user.id
    
    if complaint.status == 'resolved' or complaint.status == 'closed':
        complaint.resolved_at = datetime.now(timezone.utc)
    
    # Create notification for student if status changed
    if old_status != complaint.status:
        student = Student.query.get(complaint.student_id)
        notification = Notification(
            user_id=student.user_id,
            title=f'Complaint Update - {complaint.ticket_id}',
            message=f'Your complaint status has been updated to: {complaint.status.replace("_", " ").title()}',
            type='complaint'
        )
        db.session.add(notification)
    
    db.session.commit()
    flash('Complaint updated successfully', 'success')
    return redirect(url_for(get_staff_route('warden_complaints', 'admin_complaints')))

@app.route('/warden/attendance')
@login_required
@role_required('warden', 'admin', 'principal')
def warden_attendance():
    search = request.args.get('search', '')
    query = db.session.query(Student, User)\
        .select_from(Student)\
        .join(User, Student.user_id == User.id)
    
    if search:
        query = query.filter(User.full_name.ilike(f'%{search}%'))
    
    students = query.order_by(User.full_name).all()
    return render_template('warden/attendance.html', students=students, search=search)

# Principal Module
@app.route('/principal/dashboard')
@login_required
@role_required('principal')
def principal_dashboard():
    total_students = Student.query.count()
    total_bills = Bill.query.count()
    paid_bills = Bill.query.filter_by(paid=True).count()
    total_complaints = Complaint.query.count()
    resolved_complaints = Complaint.query.filter(Complaint.status.in_(['resolved', 'closed'])).count()
    
    # Monthly attendance summary
    thirty_days_ago = datetime.now(timezone.utc) - timedelta(days=30)
    attendance_count = AttendanceLog.query.filter(AttendanceLog.timestamp >= thirty_days_ago).count()
    
    # Financial summary
    total_revenue = db.session.query(db.func.sum(Bill.amount)).filter_by(paid=True).scalar() or 0
    pending_revenue = db.session.query(db.func.sum(Bill.amount)).filter_by(paid=False).scalar() or 0
    
    return render_template('principal/dashboard.html',
                         total_students=total_students,
                         total_bills=total_bills,
                         paid_bills=paid_bills,
                         total_complaints=total_complaints,
                         resolved_complaints=resolved_complaints,
                         attendance_count=attendance_count,
                         total_revenue=total_revenue,
                         pending_revenue=pending_revenue)

@app.route('/principal/reports')
@login_required
@role_required('principal')
def principal_reports():
    # Attendance report
    thirty_days_ago = datetime.now(timezone.utc) - timedelta(days=30)
    attendance_logs = db.session.query(AttendanceLog, Student, User)\
        .select_from(AttendanceLog)\
        .join(Student, AttendanceLog.student_id == Student.id)\
        .join(User, Student.user_id == User.id)\
        .filter(AttendanceLog.timestamp >= thirty_days_ago)\
        .order_by(AttendanceLog.timestamp.desc())\
        .limit(100).all()
    for log, _student, _user in attendance_logs:
        log.timestamp = to_ist(log.timestamp)
    
    # Billing report
    billing_summary = db.session.query(
        Bill.bill_type,
        db.func.count(Bill.id).label('count'),
        db.func.sum(Bill.amount).label('total'),
        db.func.sum(db.case((Bill.paid == True, Bill.amount), else_=0)).label('collected')
    ).group_by(Bill.bill_type).all()
    
    # Complaint summary
    complaint_summary = db.session.query(
        Complaint.category,
        db.func.count(Complaint.id).label('count'),
        db.func.sum(db.case((Complaint.status.in_(['resolved', 'closed']), 1), else_=0)).label('resolved')
    ).group_by(Complaint.category).all()
    
    return render_template('principal/reports.html',
                         attendance_logs=attendance_logs,
                         billing_summary=billing_summary,
                         complaint_summary=complaint_summary)

@app.route('/principal/attendance')
@login_required
@role_required('principal')
def principal_attendance():
    search = request.args.get('search', '')
    query = db.session.query(Student, User)\
        .select_from(Student)\
        .join(User, Student.user_id == User.id)
    
    if search:
        query = query.filter(User.full_name.ilike(f'%{search}%'))
    
    students = query.order_by(User.full_name).all()
    return render_template('principal/attendance.html', students=students, search=search)

# Admin Module  
@app.route('/admin/dashboard')
@login_required
@role_required('admin')
def admin_dashboard():
    total_users = User.query.count()
    students = User.query.filter_by(role='student').count()
    wardens = User.query.filter_by(role='warden').count()
    
    recent_users = User.query.order_by(User.created_at.desc()).limit(10).all()
    
    return render_template('admin/dashboard.html',
                         total_users=total_users,
                         students=students,
                         wardens=wardens,
                         recent_users=recent_users)

@app.route('/admin/users', methods=['GET', 'POST'])
@login_required
@role_required('admin')
def admin_users():
    if request.method == 'POST':
        user = User(
            username=request.form.get('username'),
            email=request.form.get('email'),
            full_name=request.form.get('full_name'),
            role=request.form.get('role')
        )
        user.set_password(request.form.get('password'))
        db.session.add(user)
        db.session.commit()
        flash('User created successfully', 'success')
        return redirect(url_for('admin_users'))
    
    users = User.query.order_by(User.created_at.desc()).all()
    return render_template('admin/users.html', users=users)

@app.route('/admin/users/toggle/<int:id>', methods=['POST'])
@login_required
@role_required('admin')
def toggle_user(id):
    user = User.query.get_or_404(id)
    user.is_active = not user.is_active
    db.session.commit()
    status = 'activated' if user.is_active else 'deactivated'
    flash(f'User {status} successfully', 'success')
    return redirect(url_for('admin_users'))

@app.route('/notifications')
@login_required
def notifications():
    user_notifications = Notification.query.filter_by(user_id=current_user.id).order_by(Notification.created_at.desc()).all()
    message_recipients_query = User.query.filter(User.id != current_user.id, User.is_active == True)
    if current_user.role == 'student':
        message_recipients_query = message_recipients_query.filter(User.role != 'student')
    message_recipients = message_recipients_query.order_by(User.full_name).all()
    can_broadcast_to_students = current_user.role in STAFF_ROLES

    # Mark all as read
    for notif in user_notifications:
        notif.read = True
    db.session.commit()
    return render_template(
        'notifications.html',
        notifications=user_notifications,
        message_recipients=message_recipients,
        can_broadcast_to_students=can_broadcast_to_students
    )

@app.route('/notifications/send', methods=['POST'])
@login_required
def send_staff_message():
    recipient_token = (request.form.get('recipient_id') or '').strip()
    subject = (request.form.get('subject') or '').strip()
    message = (request.form.get('message') or '').strip()

    if not recipient_token or not subject or not message:
        flash('Recipient, subject, and message are required.', 'error')
        return redirect(url_for('notifications'))

    if recipient_token == 'all_students':
        if current_user.role not in STAFF_ROLES:
            flash('You are not allowed to send broadcast messages.', 'error')
            return redirect(url_for('notifications'))

        student_recipients = User.query.filter_by(role='student', is_active=True).all()
        if not student_recipients:
            flash('No active students found for broadcast.', 'error')
            return redirect(url_for('notifications'))

        notifications = [
            Notification(
                user_id=student.id,
                sender_id=current_user.id,
                title=subject,
                message=message,
                type='message'
            )
            for student in student_recipients
        ]
        db.session.add_all(notifications)
        db.session.commit()
        flash(f'Broadcast sent successfully to {len(notifications)} students.', 'success')
        return redirect(url_for('notifications'))

    try:
        recipient_id = int(recipient_token)
    except ValueError:
        flash('Invalid recipient selected.', 'error')
        return redirect(url_for('notifications'))

    recipient = User.query.get(recipient_id)
    if not recipient or not recipient.is_active:
        flash('Invalid recipient selected.', 'error')
        return redirect(url_for('notifications'))
    if recipient.id == current_user.id:
        flash('Please select a different recipient.', 'error')
        return redirect(url_for('notifications'))
    if current_user.role == 'student' and recipient.role == 'student':
        flash('Students cannot message other students.', 'error')
        return redirect(url_for('notifications'))

    notification = Notification(
        user_id=recipient.id,
        sender_id=current_user.id,
        title=subject,
        message=message,
        type='message'
    )
    db.session.add(notification)
    db.session.commit()
    flash('Message sent successfully.', 'success')
    return redirect(url_for('notifications'))

@app.route('/notifications/reply/<int:id>', methods=['POST'])
@login_required
def reply_staff_message(id):
    source = Notification.query.get_or_404(id)
    reply_text = (request.form.get('reply_message') or '').strip()

    if not reply_text:
        flash('Reply message is required.', 'error')
        return redirect(url_for('notifications'))

    if not source.sender_id:
        flash('This notification cannot be replied to.', 'error')
        return redirect(url_for('notifications'))

    recipient = User.query.get(source.sender_id)
    if not recipient or not recipient.is_active:
        flash('Original sender is not available for reply.', 'error')
        return redirect(url_for('notifications'))
    if current_user.role == 'student' and recipient.role == 'student':
        flash('Students cannot message other students.', 'error')
        return redirect(url_for('notifications'))

    base_subject = source.title or 'Message'
    subject = base_subject if base_subject.startswith('Re:') else f'Re: {base_subject}'

    reply_notification = Notification(
        user_id=recipient.id,
        sender_id=current_user.id,
        parent_notification_id=source.id,
        title=subject,
        message=reply_text,
        type='message'
    )
    db.session.add(reply_notification)
    db.session.commit()
    flash('Reply sent successfully.', 'success')
    return redirect(url_for('notifications'))

@app.route('/api/notifications/unread-count')
@login_required
def unread_notifications():
    count = Notification.query.filter_by(user_id=current_user.id, read=False).count()
    return jsonify({'count': count})

@app.route('/rfid-scan', methods=['GET', 'POST'])
def rfid_scan():
    # If ESP sends GET request
    if request.method == 'GET':
        rfid_code = request.args.get('data_2', '').strip()
    # If mobile/app sends POST JSON
    else:
        data = request.get_json() or {}
        rfid_code = data.get('rfid_code', '').strip()

    if not rfid_code:
        return jsonify({'success': False, 'message': 'RFID code is required'}), 400

    student = Student.query.filter_by(rfid_code=rfid_code).first()
    if not student:
        return jsonify({
            'success': False,
            'message': 'Student not found. Please register this RFID code.'
        }), 404

    # Toggle status
    new_status = 'IN' if student.current_status == 'OUT' else 'OUT'
    student.current_status = new_status

    log = AttendanceLog(
        student_id=student.id,
        action=new_status,
        timestamp=datetime.now(timezone.utc)
    )
    db.session.add(log)
    db.session.commit()

    user = User.query.get(student.user_id)
    return jsonify({
        'success': True,
        'student_name': user.full_name,
        'roll_number': student.roll_number,
        'action': new_status,
        'timestamp': log.timestamp.isoformat()
    })

@app.route('/attendance/export')
@login_required
@role_required('warden', 'admin', 'principal')
def export_attendance():
    ninety_days_ago = datetime.now(timezone.utc) - timedelta(days=90)
    logs = db.session.query(AttendanceLog, Student, User)\
        .select_from(AttendanceLog)\
        .join(Student, AttendanceLog.student_id == Student.id)\
        .join(User, Student.user_id == User.id)\
        .filter(AttendanceLog.timestamp >= ninety_days_ago)\
        .order_by(AttendanceLog.timestamp.desc()).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Logs"
    
    # Headers
    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    headers = ['Timestamp', 'Student Name', 'Roll Number', 'Room Number', 'Action']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data
    for row, (log, student, user) in enumerate(logs, 2):
        ws.cell(row=row, column=1, value=to_ist(log.timestamp).strftime('%Y-%m-%d %H:%M:%S'))
        ws.cell(row=row, column=2, value=user.full_name)
        ws.cell(row=row, column=3, value=student.roll_number)
        ws.cell(row=row, column=4, value=student.room_number)
        ws.cell(row=row, column=5, value=log.action)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 10
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(output,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True,
                    download_name=f'attendance_logs_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=8001, debug=True)
